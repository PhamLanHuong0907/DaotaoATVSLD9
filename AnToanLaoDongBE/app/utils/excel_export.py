import os
import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.config import get_settings


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")


def _apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_data_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


async def export_training_list(data: dict) -> str:
    """Export training list report to Excel. Returns file path."""
    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sach huan luyen"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "DANH SÁCH HUẤN LUYỆN ATVSLĐ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    if data.get("department_name"):
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Đơn vị: {data['department_name']}"
        ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["STT", "Họ tên", "Mã NV", "Ngành nghề", "Bậc thợ",
               "Kỳ thi", "Điểm", "Xếp loại"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _apply_header_style(ws, row, len(headers))

    # Data
    for i, item in enumerate(data.get("items", []), 1):
        row += 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("full_name", ""))
        ws.cell(row=row, column=3, value=item.get("employee_id", ""))
        ws.cell(row=row, column=4, value=item.get("occupation", ""))
        ws.cell(row=row, column=5, value=item.get("skill_level", ""))
        ws.cell(row=row, column=6, value=item.get("exam_name", ""))
        ws.cell(row=row, column=7, value=item.get("score", ""))
        ws.cell(row=row, column=8, value=item.get("classification", ""))
        _apply_data_style(ws, row, len(headers))

    # Auto width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    # Save
    filename = f"training_list_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.EXPORT_DIR, filename)
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath


async def export_exam_results(data: dict) -> str:
    """Export exam results to Excel."""
    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua thi"

    ws.merge_cells("A1:G1")
    ws["A1"] = "KẾT QUẢ THI ATVSLĐ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["STT", "Họ tên", "Mã NV", "Ngành nghề", "Điểm", "Xếp loại", "Ngày thi"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _apply_header_style(ws, row, len(headers))

    for i, item in enumerate(data.get("items", []), 1):
        row += 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("full_name", ""))
        ws.cell(row=row, column=3, value=item.get("employee_id", ""))
        ws.cell(row=row, column=4, value=item.get("occupation", ""))
        ws.cell(row=row, column=5, value=item.get("score", ""))
        ws.cell(row=row, column=6, value=item.get("classification", ""))
        ws.cell(row=row, column=7, value=item.get("submitted_at", ""))
        _apply_data_style(ws, row, len(headers))

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    filename = f"exam_results_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.EXPORT_DIR, filename)
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath


async def export_individual_record(data: dict) -> str:
    """Export individual training record to Excel."""
    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ho so huan luyen"

    ws.merge_cells("A1:F1")
    ws["A1"] = "HỒ SƠ HUẤN LUYỆN CÁ NHÂN"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # User info
    user = data.get("user", {})
    ws["A3"] = "Họ tên:"
    ws["B3"] = user.get("full_name", "")
    ws["A4"] = "Mã NV:"
    ws["B4"] = user.get("employee_id", "")
    ws["A5"] = "Ngành nghề:"
    ws["B5"] = user.get("occupation", "")
    ws["A6"] = "Bậc thợ:"
    ws["B6"] = user.get("skill_level", "")
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Exam history
    headers = ["STT", "Kỳ thi", "Loại thi", "Điểm", "Xếp loại", "Ngày thi"]
    row = 8
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _apply_header_style(ws, row, len(headers))

    for i, item in enumerate(data.get("exam_history", []), 1):
        row += 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("exam_name", ""))
        ws.cell(row=row, column=3, value=item.get("exam_type", ""))
        ws.cell(row=row, column=4, value=item.get("score", ""))
        ws.cell(row=row, column=5, value=item.get("classification", ""))
        ws.cell(row=row, column=6, value=item.get("submitted_at", ""))
        _apply_data_style(ws, row, len(headers))

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    filename = f"individual_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.EXPORT_DIR, filename)
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath
