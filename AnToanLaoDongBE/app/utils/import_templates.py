"""Generate .xlsx templates for the import features.

These are written to a BytesIO so endpoints can stream them back without
touching disk.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
REQUIRED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")


def _style_header_row(ws, columns: list[dict]) -> None:
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col["name"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col.get("comment"):
            cell.comment = Comment(col["comment"], "ATVSLĐ")
        # Auto-fit-ish column width
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 18)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _add_sample_rows(ws, rows: list[list]) -> None:
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def build_user_import_template() -> bytes:
    """Build the .xlsx template for bulk-importing users."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Người dùng"

    columns = [
        {"name": "username*", "width": 16, "comment": "Bắt buộc — tên đăng nhập, không trùng"},
        {"name": "password*", "width": 14, "comment": "Bắt buộc — mật khẩu ban đầu (≥6 ký tự)"},
        {"name": "full_name*", "width": 24, "comment": "Bắt buộc — họ và tên"},
        {"name": "employee_id*", "width": 14, "comment": "Bắt buộc — mã nhân viên, không trùng"},
        {"name": "role", "width": 18, "comment": "Tùy chọn — admin / training_officer / manager / worker (mặc định: worker)"},
        {"name": "department_id", "width": 24, "comment": "Tùy chọn — ID phòng ban (lấy từ trang Phòng ban)"},
        {"name": "occupation", "width": 22, "comment": "Tùy chọn — nghề / chức danh"},
        {"name": "skill_level", "width": 12, "comment": "Tùy chọn — bậc tay nghề 1-7 (mặc định 1)"},
        {"name": "phone", "width": 14, "comment": "Tùy chọn"},
        {"name": "email", "width": 24, "comment": "Tùy chọn"},
    ]
    _style_header_row(ws, columns)

    sample_rows = [
        ["nv001", "matkhau123", "Nguyễn Văn A", "NV001", "worker",
         "", "Thợ khai thác lò", 3, "0901234567", "nva@example.com"],
        ["nv002", "matkhau123", "Trần Thị B", "NV002", "worker",
         "", "Thợ điện hầm lò", 4, "0907654321", ""],
        ["cb001", "matkhau123", "Lê Văn C", "CB001", "training_officer",
         "", "Cán bộ đào tạo", 1, "0912345678", "lvc@example.com"],
    ]
    _add_sample_rows(ws, sample_rows)

    # Instruction sheet
    info = wb.create_sheet("Hướng dẫn")
    info.column_dimensions["A"].width = 96
    info["A1"] = "HƯỚNG DẪN NHẬP NGƯỜI DÙNG TỪ EXCEL"
    info["A1"].font = Font(bold=True, size=14, color="1565C0")
    instructions = [
        "",
        "1. Cột có dấu (*) là bắt buộc.",
        "2. username và employee_id phải duy nhất — hệ thống sẽ bỏ qua dòng trùng.",
        "3. role nhận một trong các giá trị: admin, training_officer, manager, worker.",
        "   Nếu để trống, mặc định là 'worker'.",
        "4. department_id lấy từ trang Quản lý phòng ban (cột ID).",
        "5. skill_level là số nguyên từ 1 đến 7. Nếu để trống, mặc định là 1.",
        "6. Xoá các dòng mẫu trước khi upload.",
        "7. Sau khi nhập thành công, người dùng có thể đăng nhập bằng username + password ban đầu.",
        "   Khuyến nghị: yêu cầu họ đổi mật khẩu lần đầu đăng nhập.",
    ]
    for i, line in enumerate(instructions, start=2):
        info[f"A{i}"] = line

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_question_import_template() -> bytes:
    """Build the .xlsx template for bulk-importing exam questions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Câu hỏi"

    columns = [
        {"name": "content*", "width": 50, "comment": "Bắt buộc — nội dung câu hỏi"},
        {"name": "question_type*", "width": 18, "comment": "Bắt buộc — multiple_choice / true_false / scenario_based"},
        {"name": "difficulty", "width": 12, "comment": "Tùy chọn — easy / medium / hard (mặc định medium)"},
        {"name": "occupation*", "width": 22, "comment": "Bắt buộc — nghề áp dụng"},
        {"name": "skill_level*", "width": 12, "comment": "Bắt buộc — bậc tay nghề 1-7"},
        {"name": "training_group*", "width": 18, "comment": "Bắt buộc — atvsld / skill_upgrade / safety_hygiene / legal_knowledge"},
        {"name": "option_a", "width": 30, "comment": "Đáp án A (cho câu trắc nghiệm)"},
        {"name": "option_b", "width": 30, "comment": "Đáp án B"},
        {"name": "option_c", "width": 30, "comment": "Đáp án C"},
        {"name": "option_d", "width": 30, "comment": "Đáp án D"},
        {"name": "correct_label", "width": 14, "comment": "Đáp án đúng cho MCQ: A / B / C / D"},
        {"name": "correct_bool", "width": 14, "comment": "Đáp án đúng cho true_false: TRUE / FALSE"},
        {"name": "key_points", "width": 40, "comment": "Cho scenario_based: các ý chính phân tách bằng dấu |"},
        {"name": "explanation", "width": 40, "comment": "Tùy chọn — giải thích đáp án"},
        {"name": "topic_tags", "width": 24, "comment": "Tùy chọn — các tag phân tách bằng dấu phẩy"},
    ]
    _style_header_row(ws, columns)

    sample_rows = [
        [
            "Khi gặp tai nạn lao động trong hầm lò, hành động đầu tiên cần làm là gì?",
            "multiple_choice", "medium", "Thợ khai thác lò", 3, "atvsld",
            "Báo cáo cho cán bộ y tế và an toàn",
            "Tự xử lý tại chỗ",
            "Tiếp tục công việc",
            "Chờ ca sau giải quyết",
            "A", "", "",
            "Theo Luật ATVSLĐ, mọi tai nạn phải được báo cáo ngay cho cán bộ phụ trách.",
            "tai-nan, hầm-lò"
        ],
        [
            "Đèn lò an toàn phải có chứng nhận phòng nổ.",
            "true_false", "easy", "Thợ khai thác lò", 3, "atvsld",
            "", "", "", "",
            "", "TRUE", "",
            "Đèn lò sử dụng trong hầm lò bắt buộc phải đạt chứng nhận phòng nổ Ex.",
            "thiết-bị, đèn-lò"
        ],
        [
            "Mô tả các bước sơ cứu khi đồng nghiệp bị bỏng hơi nước.",
            "scenario_based", "hard", "Thợ điện hầm lò", 4, "safety_hygiene",
            "", "", "", "",
            "", "",
            "Đưa nạn nhân ra khỏi nguồn bỏng | Làm mát vết bỏng bằng nước sạch 15-20 phút | Băng vô trùng nhẹ | Đưa đến cơ sở y tế",
            "Các ý chính cần có: cách ly nguồn bỏng, làm mát, băng và chuyển viện.",
            "sơ-cứu, bỏng"
        ],
    ]
    _add_sample_rows(ws, sample_rows)

    info = wb.create_sheet("Hướng dẫn")
    info.column_dimensions["A"].width = 100
    info["A1"] = "HƯỚNG DẪN NHẬP CÂU HỎI TỪ EXCEL"
    info["A1"].font = Font(bold=True, size=14, color="1565C0")
    instructions = [
        "",
        "1. Cột có dấu (*) là bắt buộc.",
        "",
        "2. Loại câu hỏi (question_type):",
        "   • multiple_choice — câu trắc nghiệm 4 đáp án A/B/C/D.",
        "       Cần điền: option_a, option_b, option_c, option_d, correct_label (A|B|C|D).",
        "   • true_false — câu đúng/sai.",
        "       Cần điền: correct_bool (TRUE hoặc FALSE).",
        "   • scenario_based — câu tình huống tự luận.",
        "       Cần điền: key_points — các ý chính phân tách bằng dấu | (gạch đứng).",
        "       Hệ thống chấm bằng cách đếm số ý chính khớp trong câu trả lời.",
        "",
        "3. training_group nhận một trong: atvsld, skill_upgrade, safety_hygiene, legal_knowledge.",
        "",
        "4. difficulty: easy / medium / hard. Mặc định là medium nếu để trống.",
        "",
        "5. topic_tags: các tag phân tách bằng dấu phẩy hoặc dấu | (ví dụ: 'hầm-lò, sơ-cứu').",
        "",
        "6. Câu hỏi sau khi nhập sẽ ở trạng thái DRAFT — cần phê duyệt qua Hộp duyệt trước khi sử dụng.",
        "",
        "7. Xoá các dòng mẫu trước khi upload.",
    ]
    for i, line in enumerate(instructions, start=2):
        info[f"A{i}"] = line

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
