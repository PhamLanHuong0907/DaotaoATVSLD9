"""Bulk-import questions from an Excel file.

Expected columns (header row, case-insensitive):
  content, question_type, difficulty, occupation, skill_level, training_group,
  option_a, option_b, option_c, option_d, correct_label,
  correct_bool, key_points, explanation, topic_tags, created_by

Rules:
- `question_type` ∈ multiple_choice | true_false | scenario_based
- For multiple_choice: option_a..option_d are answer texts; correct_label = A/B/C/D
- For true_false: correct_bool = TRUE/FALSE/1/0/yes/no
- For scenario_based: key_points = pipe-separated key points
- topic_tags = comma- or pipe-separated
"""
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

from app.models.enums import (
    ApprovalStatus, DifficultyLevel, QuestionType, TrainingGroup,
)
from app.models.question import AnswerOption, Question


REQUIRED_COLS = ["content", "question_type", "occupation", "skill_level", "training_group"]


def _norm(h: Optional[str]) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _to_bool(v) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "đúng", "dung", "t"):
        return True
    if s in ("false", "0", "no", "n", "sai", "f"):
        return False
    return None


async def import_questions_from_xlsx(file_bytes: bytes, default_creator: str) -> dict:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"created": 0, "skipped": 0, "errors": ["Empty file"]}

    cols = {_norm(h): i for i, h in enumerate(header)}
    missing = [c for c in REQUIRED_COLS if c not in cols]
    if missing:
        return {"created": 0, "skipped": 0, "errors": [f"Missing required columns: {missing}"]}

    created = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows, start=2):
        if row is None or all(v is None for v in row):
            continue

        def get(col: str) -> Optional[str]:
            idx = cols.get(col)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return None if v is None else str(v).strip()

        try:
            content = get("content")
            qt_raw = (get("question_type") or "").lower()
            occupation = get("occupation")
            sl_raw = get("skill_level")
            tg_raw = (get("training_group") or "").lower()
            if not (content and qt_raw and occupation and sl_raw and tg_raw):
                errors.append(f"Row {line_no}: missing required field")
                skipped += 1
                continue
            try:
                qt = QuestionType(qt_raw)
            except ValueError:
                errors.append(f"Row {line_no}: invalid question_type '{qt_raw}'")
                skipped += 1
                continue
            try:
                tg = TrainingGroup(tg_raw)
            except ValueError:
                errors.append(f"Row {line_no}: invalid training_group '{tg_raw}'")
                skipped += 1
                continue
            try:
                sl = int(sl_raw)
                if not 1 <= sl <= 7:
                    raise ValueError
            except Exception:
                errors.append(f"Row {line_no}: skill_level must be 1-7")
                skipped += 1
                continue

            difficulty_raw = (get("difficulty") or "medium").lower()
            try:
                difficulty = DifficultyLevel(difficulty_raw)
            except ValueError:
                difficulty = DifficultyLevel.MEDIUM

            options: list[AnswerOption] = []
            correct_bool: Optional[bool] = None
            key_points: list[str] = []

            if qt == QuestionType.MULTIPLE_CHOICE:
                correct_label = (get("correct_label") or "").strip().upper()
                if correct_label not in ("A", "B", "C", "D"):
                    errors.append(f"Row {line_no}: correct_label must be A/B/C/D")
                    skipped += 1
                    continue
                for letter in ("A", "B", "C", "D"):
                    text = get(f"option_{letter.lower()}")
                    if text:
                        options.append(AnswerOption(
                            label=letter, text=text, is_correct=(letter == correct_label),
                        ))
                if not any(o.is_correct for o in options):
                    errors.append(f"Row {line_no}: no option matches correct_label")
                    skipped += 1
                    continue
            elif qt == QuestionType.TRUE_FALSE:
                cb = _to_bool(get("correct_bool"))
                if cb is None:
                    errors.append(f"Row {line_no}: correct_bool must be TRUE/FALSE")
                    skipped += 1
                    continue
                correct_bool = cb
            elif qt == QuestionType.SCENARIO_BASED:
                kps_raw = get("key_points") or ""
                key_points = [k.strip() for k in kps_raw.split("|") if k.strip()]
                if not key_points:
                    errors.append(f"Row {line_no}: scenario question needs key_points")
                    skipped += 1
                    continue

            tags_raw = get("topic_tags") or ""
            tags = [t.strip() for t in tags_raw.replace("|", ",").split(",") if t.strip()]

            q = Question(
                content=content,
                question_type=qt,
                difficulty=difficulty,
                options=options,
                correct_answer_bool=correct_bool,
                expected_key_points=key_points,
                explanation=get("explanation"),
                occupation=occupation,
                skill_level=sl,
                training_group=tg,
                topic_tags=tags,
                created_by=get("created_by") or default_creator,
                status=ApprovalStatus.DRAFT,
            )
            await q.insert()
            created += 1
        except Exception as e:
            errors.append(f"Row {line_no}: {e}")
            skipped += 1

    return {"created": created, "skipped": skipped, "errors": errors}
