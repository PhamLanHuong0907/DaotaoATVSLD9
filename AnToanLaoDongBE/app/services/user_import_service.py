"""Bulk-import users from an Excel (.xlsx) file.

Expected columns (header row, case-insensitive):
  username, password, full_name, employee_id, role, department_id,
  occupation, skill_level, phone, email
"""
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

from app.models.user import User
from app.models.enums import UserRole
from app.utils.security import hash_password


REQUIRED_COLS = ["username", "password", "full_name", "employee_id"]
ALL_COLS = REQUIRED_COLS + [
    "role", "department_id", "occupation", "skill_level", "phone", "email",
]


def _normalize_header(h: Optional[str]) -> str:
    return (h or "").strip().lower().replace(" ", "_")


async def import_users_from_xlsx(file_bytes: bytes) -> dict:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"created": 0, "skipped": 0, "errors": ["Empty file"]}

    col_index = {_normalize_header(h): i for i, h in enumerate(header)}
    missing = [c for c in REQUIRED_COLS if c not in col_index]
    if missing:
        return {"created": 0, "skipped": 0, "errors": [f"Missing required columns: {missing}"]}

    created = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows, start=2):
        if row is None or all(v is None for v in row):
            continue

        def get(col: str) -> Optional[str]:
            idx = col_index.get(col)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return None if v is None else str(v).strip()

        try:
            username = get("username")
            password = get("password")
            full_name = get("full_name")
            employee_id = get("employee_id")
            if not (username and password and full_name and employee_id):
                errors.append(f"Row {line_no}: missing required field")
                skipped += 1
                continue

            role_raw = (get("role") or "worker").lower()
            try:
                role = UserRole(role_raw)
            except ValueError:
                errors.append(f"Row {line_no}: invalid role '{role_raw}'")
                skipped += 1
                continue

            skill_level_raw = get("skill_level")
            skill_level = int(skill_level_raw) if skill_level_raw else 1
            if not 1 <= skill_level <= 7:
                errors.append(f"Row {line_no}: skill_level must be 1-7")
                skipped += 1
                continue

            # Dedup
            if await User.find_one(User.username == username):
                errors.append(f"Row {line_no}: username '{username}' exists")
                skipped += 1
                continue
            if await User.find_one(User.employee_id == employee_id):
                errors.append(f"Row {line_no}: employee_id '{employee_id}' exists")
                skipped += 1
                continue

            user = User(
                username=username,
                password_hash=hash_password(password),
                full_name=full_name,
                employee_id=employee_id,
                role=role,
                department_id=get("department_id"),
                occupation=get("occupation"),
                skill_level=skill_level,
                phone=get("phone"),
                email=get("email"),
            )
            await user.insert()
            created += 1
        except Exception as e:
            errors.append(f"Row {line_no}: {e}")
            skipped += 1

    return {"created": created, "skipped": skipped, "errors": errors}
